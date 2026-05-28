"""Refresh bgk_auctions from BGK's "Baza obligacji" XLSX.

Scrapes the Statystyka index page for the current XLSX URL, downloads
the file, parses one row per issue event, and upserts to bgk_auctions
keyed on (issue_date, isin).

XLSX schema (BGK ships bilingual headers on one row, e.g. "Seria\n Series"
or "Typ transakcji* \nTransaction type*" - we normalize to the Polish
prefix and match candidates against that):
    Seria, Data emisji, Data wykupu, ISIN, Typ transakcji,
    Lata do wykupu, Oprocentowanie, Kupon, Waluta, Wartość emisji,
    Cena, Rentowność

Coupon type values are also bilingual ("stałe/fixed", "zmienne/floating");
we strip the English half on read.

Percent fields (Kupon / Cena / Rentowność) are stored in XLSX as decimal
fractions (0.0575 = 5.75%) and we convert to percent before upsert.

Idempotent via PRIMARY KEY (issue_date, isin); reruns overwrite.
"""

from __future__ import annotations

import sys
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from lib.bgk_xlsx import download_xlsx, find_xlsx_url  # noqa: E402
from lib.supabase import upsert  # noqa: E402


# Column-header lookup. We compare against the *normalized Polish prefix*
# of each header (see _normalize_header) so the same candidate works
# whether BGK ships "Seria", "Seria\n Series", "Seria*" etc.
_COLUMN_CANDIDATES: dict[str, list[str]] = {
    "series":            ["Seria"],
    "issue_date":        ["Data emisji"],
    "maturity_date":     ["Data wykupu"],
    "isin":              ["ISIN"],
    "program":           ["Typ transakcji"],
    "years_to_maturity": ["Lata do wykupu"],
    "coupon_kind":       ["Oprocentowanie"],
    "coupon_pct":        ["Kupon"],
    "currency":          ["Waluta"],
    "issue_amount":      ["Wartość emisji", "Wartosc emisji"],
    "price_pct":         ["Cena"],
    "yield_pct":         ["Rentowność", "Rentownosc"],
}


def _normalize_header(raw) -> str:
    """Reduce a BGK bilingual header to its Polish prefix.

    Examples:
        "Seria\\n Series"              -> "Seria"
        "Typ transakcji* \\nTransaction type*" -> "Typ transakcji"
        "Kupon  \\nCoupon"             -> "Kupon"
    """
    if raw is None:
        return ""
    s = str(raw)
    # The Polish half always comes first, separated from English by newline.
    if "\n" in s:
        s = s.split("\n", 1)[0]
    # Some headers carry a trailing '*' footnote marker.
    return s.strip().rstrip("*").strip()


def _normalize_coupon_kind(raw) -> str | None:
    """'stałe/fixed' -> 'stałe', 'zmienne/floating' -> 'zmienne'."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    return s.split("/", 1)[0].strip()


def _to_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        # Possible string formats: "DD.MM.YYYY" or "YYYY-MM-DD"
        for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                pass
    return None


def _to_float(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _decimal_to_pct(value) -> float | None:
    """BGK XLSX stores percents as decimal (0.0575 = 5.75%). Multiply by 100."""
    f = _to_float(value)
    return f * 100.0 if f is not None else None


def _to_str(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _build_header_index(headers: list) -> dict[str, int]:
    """Map each logical field to the column index that matched a candidate."""
    idx: dict[str, int] = {}
    norm_headers = [_normalize_header(h) for h in headers]
    for field, candidates in _COLUMN_CANDIDATES.items():
        for cand in candidates:
            if cand in norm_headers:
                idx[field] = norm_headers.index(cand)
                break
    missing = set(_COLUMN_CANDIDATES) - set(idx)
    if missing:
        raise RuntimeError(
            f"BGK XLSX missing expected columns: {sorted(missing)}. "
            f"Got normalized headers: {norm_headers}"
        )
    return idx


def _find_data_sheet(wb) -> tuple[str, int]:
    """Pick the sheet whose first row contains the expected BGK headers.

    Returns (sheet_name, header_row_index_1based). Header may not be on row 1
    if the sheet has a banner / title row at the top.
    """
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Scan the first 5 rows for one whose normalized cells include 'Seria' AND 'ISIN'.
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
            normalized = {_normalize_header(c) for c in row if c is not None}
            if "Seria" in normalized and "ISIN" in normalized:
                return sheet, row_idx
    raise RuntimeError(
        "Could not find BGK data sheet (no sheet has both 'Seria' and 'ISIN' "
        f"headers in the first 5 rows). Sheets present: {wb.sheetnames}"
    )


def parse_bgk(xlsx_bytes: BytesIO, source_url: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_bytes, read_only=True, data_only=True)
    sheet_name, header_row = _find_data_sheet(wb)
    ws = wb[sheet_name]

    headers = list(next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)))
    col = _build_header_index(headers)

    out: list[dict] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        # Skip blank rows (BGK sometimes pads with empties after the last data row).
        if all(c is None or c == "" for c in row):
            continue

        isin = _to_str(row[col["isin"]])
        issue_date = _to_date(row[col["issue_date"]])
        if not isin or not issue_date:
            continue  # malformed row - skip silently

        out.append({
            "issue_date":        issue_date.isoformat(),
            "isin":              isin,
            "series":            _to_str(row[col["series"]]) or "",
            "maturity_date":     _to_date(row[col["maturity_date"]]).isoformat()
                                 if _to_date(row[col["maturity_date"]]) else None,
            "program":           _to_str(row[col["program"]]),
            "years_to_maturity": int(_to_float(row[col["years_to_maturity"]]))
                                 if _to_float(row[col["years_to_maturity"]]) is not None else None,
            "coupon_kind":       _normalize_coupon_kind(row[col["coupon_kind"]]),
            "coupon_pct":        _decimal_to_pct(row[col["coupon_pct"]]),
            "currency":          (_to_str(row[col["currency"]]) or "PLN")[:3],
            "issue_amount":      _to_float(row[col["issue_amount"]]),
            "price_pct":         _decimal_to_pct(row[col["price_pct"]]),
            "yield_pct":         _decimal_to_pct(row[col["yield_pct"]]),
            "source":            "bgk_xlsx",
            "source_url":        source_url,
        })

    return out


def _print_summary(rows: list[dict]) -> None:
    if not rows:
        print("  -> 0 rows parsed", flush=True)
        return
    by_program: dict[str, int] = {}
    by_currency: dict[str, int] = {}
    for r in rows:
        p = r["program"] or "(null)"
        c = r["currency"] or "(null)"
        by_program[p] = by_program.get(p, 0) + 1
        by_currency[c] = by_currency.get(c, 0) + 1
    print(f"  -> {len(rows)} rows parsed", flush=True)
    print(f"  -> by program: {dict(sorted(by_program.items(), key=lambda x: -x[1]))}", flush=True)
    print(f"  -> by currency: {by_currency}", flush=True)
    recent = sorted(rows, key=lambda r: r["issue_date"], reverse=True)[:5]
    print("  -> 5 most recent issuances:", flush=True)
    for r in recent:
        amt = r["issue_amount"] or 0
        yld = r["yield_pct"]
        yld_s = f"{yld:.3f}%" if yld is not None else "  n/a"
        print(
            f"     {r['issue_date']}  {r['series']:<10} {r['isin']:<12}  "
            f"mat={r['maturity_date']}  amt={amt/1e6:>9.1f}m {r['currency']}  "
            f"yld={yld_s}",
            flush=True,
        )


def main() -> None:
    print("[1/4] Locating BGK Baza_obligacji XLSX URL...", flush=True)
    url, snapshot = find_xlsx_url()
    print(f"  -> {url}", flush=True)
    print(f"  -> snapshot date: {snapshot.isoformat()}", flush=True)

    print("[2/4] Downloading XLSX...", flush=True)
    xlsx = download_xlsx(url)
    print(f"  -> {xlsx.getbuffer().nbytes / 1024:.0f} KB", flush=True)

    print("[3/4] Parsing 'Baza obligacji' sheet...", flush=True)
    rows = parse_bgk(xlsx, url)
    _print_summary(rows)

    print("[4/4] Upserting to bgk_auctions...", flush=True)
    posted = upsert(
        "bgk_auctions",
        rows,
        on_conflict="issue_date,isin",
        batch_size=1000,
    )
    print(f"  -> {posted} rows posted", flush=True)
    print("Done.", flush=True)


if __name__ == "__main__":
    main()
